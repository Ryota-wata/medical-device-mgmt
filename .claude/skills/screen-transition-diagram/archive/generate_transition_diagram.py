#!/usr/bin/env python3
"""
画面遷移図生成スクリプト
スクリーンショットを配置し、直線・直角線で遷移を描画したExcelを生成
"""

import json
import os
import sys
import math
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import argparse

# 定数
CELL_WIDTH_PX = 64
CELL_HEIGHT_PX = 20
THUMBNAIL_WIDTH = 200
GRID_MARGIN = 80  # グリッド間マージンを広げる
ARROW_COLOR = (220, 20, 60)  # 赤色で目立たせる
ARROW_WIDTH = 4
LABEL_FONT_SIZE = 11


def load_transitions(json_path):
    """遷移定義JSONを読み込み"""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_screenshot(screenshots_dir, filename, scale=0.15):
    """スクリーンショットを読み込んでリサイズ"""
    path = Path(screenshots_dir) / filename
    if not path.exists():
        print(f"警告: スクリーンショットが見つかりません: {path}")
        img = Image.new('RGB', (THUMBNAIL_WIDTH, int(THUMBNAIL_WIDTH * 0.6)), (200, 200, 200))
        draw = ImageDraw.Draw(img)
        draw.text((10, 10), filename, fill=(100, 100, 100))
        return img

    img = Image.open(path)
    new_width = int(img.width * scale)
    new_height = int(img.height * scale)
    return img.resize((new_width, new_height), Image.LANCZOS)


def get_connection_points(from_screen, to_screen, screen_positions, thumbnail_sizes):
    """
    2つの画面間の接続点を計算
    screen_positions: {screen_id: (x, y)} 実際の描画位置
    """
    from_id = from_screen['id']
    to_id = to_screen['id']

    from_x, from_y = screen_positions[from_id]
    to_x, to_y = screen_positions[to_id]

    from_w, from_h = thumbnail_sizes.get(from_id, (THUMBNAIL_WIDTH, 150))
    to_w, to_h = thumbnail_sizes.get(to_id, (THUMBNAIL_WIDTH, 150))

    from_row = from_screen['position']['row']
    from_col = from_screen['position']['col']
    to_row = to_screen['position']['row']
    to_col = to_screen['position']['col']

    # 接続タイプと始点・終点を決定
    if from_row == to_row:  # 同じ行（水平）
        if from_col < to_col:  # 右へ
            return (from_x + from_w + 5, from_y + from_h // 2,
                    to_x - 5, to_y + to_h // 2, 'horizontal')
        else:  # 左へ
            return (from_x - 5, from_y + from_h // 2,
                    to_x + to_w + 5, to_y + to_h // 2, 'horizontal')

    elif from_col == to_col:  # 同じ列（垂直）
        if from_row < to_row:  # 下へ
            return (from_x + from_w // 2, from_y + from_h + 5,
                    to_x + to_w // 2, to_y - 5, 'vertical')
        else:  # 上へ
            return (from_x + from_w // 2, from_y - 5,
                    to_x + to_w // 2, to_y + to_h + 5, 'vertical')

    else:  # 直角線（エルボー）
        if from_row < to_row:  # 下方向
            return (from_x + from_w // 2, from_y + from_h + 5,
                    to_x + to_w // 2, to_y - 5, 'elbow_down')
        else:  # 上方向
            return (from_x + from_w // 2, from_y - 5,
                    to_x + to_w // 2, to_y + to_h + 5, 'elbow_up')


def draw_arrow(draw, x1, y1, x2, y2, color=ARROW_COLOR, width=ARROW_WIDTH):
    """矢印を描画"""
    # 線を描画
    draw.line([(x1, y1), (x2, y2)], fill=color, width=width)

    # 矢印の先端を描画
    angle = math.atan2(y2 - y1, x2 - x1)
    arrow_length = 18
    arrow_angle = math.pi / 6

    ax1 = x2 - arrow_length * math.cos(angle - arrow_angle)
    ay1 = y2 - arrow_length * math.sin(angle - arrow_angle)
    ax2 = x2 - arrow_length * math.cos(angle + arrow_angle)
    ay2 = y2 - arrow_length * math.sin(angle + arrow_angle)

    # 三角形の矢印
    draw.polygon([(x2, y2), (ax1, ay1), (ax2, ay2)], fill=color)


def draw_elbow_arrow(draw, x1, y1, x2, y2, elbow_type, color=ARROW_COLOR, width=ARROW_WIDTH):
    """直角線の矢印を描画"""
    if elbow_type == 'elbow_down':
        # 下へ行って横に曲がる
        mid_y = (y1 + y2) // 2
        draw.line([(x1, y1), (x1, mid_y)], fill=color, width=width)
        draw.line([(x1, mid_y), (x2, mid_y)], fill=color, width=width)
        draw.line([(x2, mid_y), (x2, y2)], fill=color, width=width)

        # 矢印
        arrow_length = 18
        arrow_angle = math.pi / 6
        angle = math.pi / 2  # 下向き

        ax1 = x2 - arrow_length * math.cos(angle - arrow_angle)
        ay1 = y2 - arrow_length * math.sin(angle - arrow_angle)
        ax2 = x2 - arrow_length * math.cos(angle + arrow_angle)
        ay2 = y2 - arrow_length * math.sin(angle + arrow_angle)
        draw.polygon([(x2, y2), (ax1, ay1), (ax2, ay2)], fill=color)
    else:
        # 上へ行って横に曲がる
        mid_y = (y1 + y2) // 2
        draw.line([(x1, y1), (x1, mid_y)], fill=color, width=width)
        draw.line([(x1, mid_y), (x2, mid_y)], fill=color, width=width)
        draw.line([(x2, mid_y), (x2, y2)], fill=color, width=width)

        # 矢印
        arrow_length = 18
        arrow_angle = math.pi / 6
        angle = -math.pi / 2  # 上向き

        ax1 = x2 - arrow_length * math.cos(angle - arrow_angle)
        ay1 = y2 - arrow_length * math.sin(angle - arrow_angle)
        ax2 = x2 - arrow_length * math.cos(angle + arrow_angle)
        ay2 = y2 - arrow_length * math.sin(angle + arrow_angle)
        draw.polygon([(x2, y2), (ax1, ay1), (ax2, ay2)], fill=color)


def generate_diagram_image(transitions_data, screenshots_dir, scale=0.15):
    """遷移図画像を生成"""
    screens = {s['id']: s for s in transitions_data['screens']}
    transitions = transitions_data.get('transitions', [])

    # サムネイルを読み込み
    thumbnails = {}
    thumbnail_sizes = {}
    for screen in transitions_data['screens']:
        img = load_screenshot(screenshots_dir, screen['screenshot'], scale)
        thumbnails[screen['id']] = img
        thumbnail_sizes[screen['id']] = img.size

    # キャンバスサイズを計算
    max_row = max(s['position']['row'] for s in transitions_data['screens'])
    max_col = max(s['position']['col'] for s in transitions_data['screens'])

    max_thumb_w = max(s[0] for s in thumbnail_sizes.values()) if thumbnail_sizes else THUMBNAIL_WIDTH
    max_thumb_h = max(s[1] for s in thumbnail_sizes.values()) if thumbnail_sizes else 150

    grid_spacing_x = max_thumb_w + GRID_MARGIN
    grid_spacing_y = max_thumb_h + GRID_MARGIN + 40

    # グループラベル用の左マージン
    group_label_width = 200
    margin = 60
    canvas_width = group_label_width + margin + max_col * grid_spacing_x + 50
    canvas_height = margin * 2 + max_row * grid_spacing_y

    # キャンバス作成
    canvas = Image.new('RGB', (canvas_width, canvas_height), (255, 255, 255))
    draw = ImageDraw.Draw(canvas)

    # フォント
    try:
        font = ImageFont.truetype("/System/Library/Fonts/ヒラギノ角ゴシック W4.ttc", LABEL_FONT_SIZE)
        group_font = ImageFont.truetype("/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc", 13)
    except:
        font = ImageFont.load_default()
        group_font = font

    # 1. グループラベルを描画
    row_groups = {}
    for screen in transitions_data['screens']:
        row = screen['position']['row']
        if row not in row_groups:
            row_groups[row] = screen.get('group', '')

    for row, group_name in row_groups.items():
        if group_name and group_name != 'メイン遷移':
            y = margin + (row - 1) * grid_spacing_y + 50
            draw.rectangle([10, y - 5, group_label_width - 10, y + 22],
                          fill=(230, 240, 255), outline=(70, 130, 180), width=2)
            draw.text((15, y), group_name, fill=(25, 25, 112), font=group_font)

    # 2. 画面を配置（先にサムネイルを置く）+ 位置を記録
    effective_margin = group_label_width + margin
    screen_positions = {}  # {screen_id: (thumb_x, thumb_y)}

    for screen in transitions_data['screens']:
        row = screen['position']['row'] - 1
        col = screen['position']['col'] - 1

        x = effective_margin + col * grid_spacing_x
        y = margin + row * grid_spacing_y

        # 画面名ラベル
        label = screen['name']
        bbox = draw.textbbox((0, 0), label, font=font)
        text_w = bbox[2] - bbox[0]
        thumb_w, thumb_h = thumbnail_sizes[screen['id']]
        label_x = x + (thumb_w - text_w) // 2
        draw.text((label_x, y), label, fill=(30, 30, 30), font=font)

        # サムネイルの位置を記録（ラベル下）
        thumb_x = x
        thumb_y = y + 25
        screen_positions[screen['id']] = (thumb_x, thumb_y)

        # サムネイル枠線
        draw.rectangle(
            [thumb_x - 3, thumb_y - 3, thumb_x + thumb_w + 3, thumb_y + thumb_h + 3],
            outline=(100, 100, 100), width=2
        )

        # サムネイル貼り付け
        canvas.paste(thumbnails[screen['id']], (thumb_x, thumb_y))

    # 3. 遷移線を描画（画面の上に描画）
    for trans in transitions:
        from_screen = screens.get(trans['from'])
        to_screen = screens.get(trans['to'])

        if not from_screen or not to_screen:
            print(f"警告: 遷移が見つかりません: {trans['from']} -> {trans['to']}")
            continue

        x1, y1, x2, y2, conn_type = get_connection_points(
            from_screen, to_screen, screen_positions, thumbnail_sizes
        )

        print(f"描画: {trans['from']} -> {trans['to']} ({x1},{y1}) -> ({x2},{y2}) [{conn_type}]")

        if conn_type in ('horizontal', 'vertical'):
            draw_arrow(draw, x1, y1, x2, y2)
        else:
            draw_elbow_arrow(draw, x1, y1, x2, y2, conn_type)

        # ラベルを描画
        if trans.get('label'):
            mid_x = (x1 + x2) // 2
            mid_y = (y1 + y2) // 2
            label_text = trans['label']
            bbox = draw.textbbox((0, 0), label_text, font=font)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]
            # 背景
            draw.rectangle(
                [mid_x - text_w // 2 - 4, mid_y - text_h // 2 - 3,
                 mid_x + text_w // 2 + 4, mid_y + text_h // 2 + 3],
                fill=(255, 255, 220), outline=(200, 180, 0)
            )
            draw.text((mid_x - text_w // 2, mid_y - text_h // 2),
                      label_text, fill=(80, 80, 80), font=font)

    return canvas


def create_excel_with_diagram(diagram_image, output_path, title="画面遷移図"):
    """Excelファイルに遷移図画像を貼り付け"""
    wb = Workbook()
    ws = wb.active
    ws.title = title

    temp_image_path = "/tmp/transition_diagram_temp.png"
    diagram_image.save(temp_image_path)

    img = XLImage(temp_image_path)
    ws.add_image(img, 'A1')

    ws.column_dimensions['A'].width = diagram_image.width / CELL_WIDTH_PX + 5

    wb.save(output_path)
    print(f"画面遷移図を保存しました: {output_path}")

    os.remove(temp_image_path)


def main():
    parser = argparse.ArgumentParser(description='画面遷移図を生成')
    parser.add_argument('--transitions', default=None, help='遷移定義JSONパス')
    parser.add_argument('--screenshots', default='/Users/watanaberyouta/Desktop/画面設計書/screenshots',
                        help='スクリーンショットディレクトリ')
    parser.add_argument('--output', default='/Users/watanaberyouta/Desktop/画面設計書/画面遷移図.xlsx',
                        help='出力Excelパス')
    parser.add_argument('--scale', type=float, default=0.10, help='サムネイルスケール')

    args = parser.parse_args()

    transitions_path = args.transitions or '/Users/watanaberyouta/Desktop/画面設計書/transitions/transitions.json'
    transitions_data = load_transitions(transitions_path)

    print(f"画面数: {len(transitions_data['screens'])}")
    print(f"遷移数: {len(transitions_data.get('transitions', []))}")

    print("遷移図を生成中...")
    diagram = generate_diagram_image(transitions_data, args.screenshots, args.scale)

    create_excel_with_diagram(diagram, args.output)


if __name__ == '__main__':
    main()
