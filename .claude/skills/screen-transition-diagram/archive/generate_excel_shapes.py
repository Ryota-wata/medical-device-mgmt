#!/usr/bin/env python3
"""
Excelネイティブ図形で画面遷移図を生成
openpyxlのShapeを使用して個別オブジェクトとして配置
"""

import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


def load_transitions(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def generate_excel_diagram(transitions_data, output_path):
    """Excelネイティブ図形で画面遷移図を生成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "画面遷移図"

    # 列幅と行高さを調整
    for i in range(1, 20):
        ws.column_dimensions[chr(64 + i)].width = 15
    for i in range(1, 50):
        ws.row_dimensions[i].height = 30

    screens = {s['id']: s for s in transitions_data['screens']}
    transitions = transitions_data.get('transitions', [])

    # 画面ボックスの位置を記録
    box_positions = {}

    # 画面ごとにテキストをセルに配置（図形の代わり）
    for screen in transitions_data['screens']:
        row = screen['position']['row']
        col = screen['position']['col']

        # セル位置を計算（2行おき、2列おきに配置）
        cell_row = row * 3
        cell_col = col * 2 + 1

        # セルに画面名を書き込み
        cell = ws.cell(row=cell_row, column=cell_col, value=screen['name'])
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # セルの背景色
        cell.fill = PatternFill(start_color="E0F0FF", end_color="E0F0FF", fill_type="solid")
        cell.border = Border(
            left=Side(style='medium', color='4682B4'),
            right=Side(style='medium', color='4682B4'),
            top=Side(style='medium', color='4682B4'),
            bottom=Side(style='medium', color='4682B4')
        )

        box_positions[screen['id']] = (cell_row, cell_col)

    # 遷移を別セルに「→」で表現
    for trans in transitions:
        from_screen = screens.get(trans['from'])
        to_screen = screens.get(trans['to'])

        if not from_screen or not to_screen:
            continue

        from_row, from_col = box_positions[from_screen['id']]
        to_row, to_col = box_positions[to_screen['id']]

        # 矢印を配置する位置を計算
        if from_row == to_row:  # 水平
            arrow_row = from_row
            arrow_col = (from_col + to_col) // 2
            if from_col < to_col:
                arrow = "→"
            else:
                arrow = "←"
        elif from_col == to_col:  # 垂直
            arrow_row = (from_row + to_row) // 2
            arrow_col = from_col
            if from_row < to_row:
                arrow = "↓"
            else:
                arrow = "↑"
        else:  # 斜め
            arrow_row = (from_row + to_row) // 2
            arrow_col = (from_col + to_col) // 2
            if from_row < to_row and from_col < to_col:
                arrow = "↘"
            elif from_row < to_row and from_col > to_col:
                arrow = "↙"
            elif from_row > to_row and from_col < to_col:
                arrow = "↗"
            else:
                arrow = "↖"

        # 矢印セルに書き込み
        arrow_cell = ws.cell(row=arrow_row, column=arrow_col, value=arrow)
        arrow_cell.alignment = Alignment(horizontal='center', vertical='center')
        arrow_cell.font = Font(size=16, color="DC143C", bold=True)

        # ラベルがあれば追加
        if trans.get('label'):
            label_cell = ws.cell(row=arrow_row + 1, column=arrow_col, value=trans['label'])
            label_cell.alignment = Alignment(horizontal='center', vertical='center')
            label_cell.font = Font(size=9, color="666666")

    wb.save(output_path)
    print(f"画面遷移図を保存しました: {output_path}")


def main():
    transitions_path = '/Users/watanaberyouta/Desktop/画面設計書/transitions/transitions.json'
    output_path = '/Users/watanaberyouta/Desktop/画面設計書/画面遷移図_shapes.xlsx'

    transitions_data = load_transitions(transitions_path)

    print(f"画面数: {len(transitions_data['screens'])}")
    print(f"遷移数: {len(transitions_data.get('transitions', []))}")

    generate_excel_diagram(transitions_data, output_path)


if __name__ == '__main__':
    main()
