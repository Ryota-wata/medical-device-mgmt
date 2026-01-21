#!/usr/bin/env python3
"""
画面設計書Excel生成スクリプト（全画面統合版）

使用方法:
    python3 generate_excel.py --output-dir "/path/to/output" --elements-dir "/path/to/elements"

依存関係:
    pip install openpyxl
"""

import argparse
import os
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("openpyxlがインストールされていません。以下のコマンドでインストールしてください：")
    print("  pip install openpyxl")
    exit(1)


# スタイル定義
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


def parse_elements_md(file_path: str) -> list[dict]:
    """
    elements.mdファイルをパースして要素リストを返す

    期待フォーマット:
    | No | 要素名 | 要素種別 | 必須 | 機能説明 | 初期値 | バリデーション | 備考 |
    """
    elements = []

    if not os.path.exists(file_path):
        return elements

    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')
    table_lines = [line for line in lines if line.strip().startswith('|')]

    if len(table_lines) < 3:
        return elements

    for line in table_lines[2:]:
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        if len(cells) >= 5:
            elements.append({
                'no': cells[0],
                'element_name': cells[1],
                'element_type': cells[2],
                'required': cells[3] if len(cells) > 3 else '',
                'function': cells[4] if len(cells) > 4 else '',
                'initial_value': cells[5] if len(cells) > 5 else '',
                'validation': cells[6] if len(cells) > 6 else '',
                'remarks': cells[7] if len(cells) > 7 else ''
            })

    return elements


def apply_header_style(ws, row, columns):
    """ヘッダー行にスタイルを適用"""
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER_ALIGN


def create_index_sheet(wb) -> any:
    """目次シートを作成"""
    ws = wb.active
    ws.title = "目次"

    headers = ["No", "画面名", "画面モックURL", "概要", "備考"]
    apply_header_style(ws, 1, headers)

    col_widths = [6, 20, 40, 30, 20]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    return ws


def create_number_image(number: int, output_path: str, size: int = 30):
    """番号マーカー画像を生成"""
    from PIL import Image, ImageDraw, ImageFont

    # 円形の背景を作成
    img = Image.new('RGBA', (size, size), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    # 赤い円を描画
    draw.ellipse([0, 0, size-1, size-1], fill=(231, 76, 60, 255))

    # 番号テキストを描画
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", int(size * 0.5))
    except:
        font = ImageFont.load_default()

    text = str(number)
    bbox = draw.textbbox((0, 0), text, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    x = (size - text_width) // 2
    y = (size - text_height) // 2 - 2
    draw.text((x, y), text, fill=(255, 255, 255, 255), font=font)

    img.save(output_path, 'PNG')
    return output_path


def draw_markers_on_screenshot(screenshot_path: str, positions: dict, output_path: str, marker_size: int = 30):
    """
    スクリーンショット画像に直接マーカーを描画する

    Args:
        screenshot_path: 元のスクリーンショットパス
        positions: {no: (x, y)} の形式の位置情報
        output_path: マーカー付き画像の出力パス
        marker_size: マーカーサイズ
    """
    from PIL import Image, ImageDraw, ImageFont

    # スクリーンショットを開く
    screenshot = Image.open(screenshot_path).convert('RGBA')
    draw = ImageDraw.Draw(screenshot)

    # フォント設定
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", int(marker_size * 0.5))
    except:
        font = ImageFont.load_default()

    # 各マーカーを描画
    for no, (x, y) in positions.items():
        # 赤い円を描画
        draw.ellipse(
            [x, y, x + marker_size - 1, y + marker_size - 1],
            fill=(231, 76, 60, 255)
        )

        # 番号テキストを描画
        text = str(no)
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        text_x = x + (marker_size - text_width) // 2
        text_y = y + (marker_size - text_height) // 2 - 2
        draw.text((text_x, text_y), text, fill=(255, 255, 255, 255), font=font)

    # 保存
    screenshot.save(output_path, 'PNG')
    return output_path


def create_number_sheet(wb, temp_dir: str = None):
    """No.画像シートを作成（1-30の番号マーカー画像）"""
    import tempfile

    ws = wb.create_sheet(title="No.画像")

    # 説明
    ws.cell(row=1, column=1, value="スクリーンショットに貼り付ける要素No.画像")
    ws.merge_cells('A1:J1')
    ws.cell(row=2, column=1, value="※ 必要な番号をコピーしてスクリーンショット上に配置してください")
    ws.merge_cells('A2:J2')

    # 一時ディレクトリ
    if temp_dir is None:
        temp_dir = tempfile.mkdtemp()

    # 列幅と行高さを設定
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 5

    # 1-30の番号画像を10列×3行で配置
    start_row = 4
    marker_size = 30

    for row in range(start_row, start_row + 3):
        ws.row_dimensions[row].height = 30

    for i in range(1, 31):
        row = start_row + ((i - 1) // 10)
        col = ((i - 1) % 10) + 1

        # 番号画像を生成
        img_path = os.path.join(temp_dir, f"num_{i}.png")
        create_number_image(i, img_path, marker_size)

        # 画像を挿入
        img = XLImage(img_path)
        img.width = marker_size
        img.height = marker_size
        cell_ref = f"{get_column_letter(col)}{row}"
        ws.add_image(img, cell_ref)

    return ws


def parse_positions_md(file_path: str) -> dict:
    """
    要素位置定義ファイルをパースして位置情報を返す

    期待フォーマット:
    | No | デバイス | X | Y |
    """
    positions = {}

    if not os.path.exists(file_path):
        return positions

    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')
    table_lines = [line for line in lines if line.strip().startswith('|')]

    if len(table_lines) < 3:
        return positions

    for line in table_lines[2:]:
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        if len(cells) >= 4:
            try:
                no = int(cells[0])
                device = cells[1]
                x = int(cells[2])
                y = int(cells[3])
                if device not in positions:
                    positions[device] = {}
                positions[device][no] = (x, y)
            except (ValueError, IndexError):
                continue

    return positions


def create_multi_device_sheet(wb, screen_name: str, elements: list[dict], screenshots_dir: str, devices: list[str] = None, positions_dir: str = None):
    """
    複数デバイス対応の要素シートを作成
    レイアウト: 上から PC, タブレット, スマホ のスクリーンショットを縦に配置、右に要素一覧

    Args:
        wb: Workbook
        screen_name: 画面名
        elements: 要素リスト
        screenshots_dir: スクリーンショットディレクトリ
        devices: デバイスリスト ['PC', 'タブレット', 'スマホ']
        positions_dir: 要素位置定義ファイルのディレクトリ
    """
    import tempfile

    if devices is None:
        devices = ['PC', 'タブレット', 'スマホ']

    ws = wb.create_sheet(title=screen_name)

    # 要素位置を読み込み
    element_positions = {}
    if positions_dir:
        positions_file = os.path.join(positions_dir, f"{screen_name}_positions.md")
        element_positions = parse_positions_md(positions_file)
        if element_positions:
            print(f"  要素位置定義を読み込み: {positions_file}")

    # 列幅設定
    col_width = 2.5
    pixels_per_col = 17
    row_height_px = 15
    marker_size = 24

    # 一時ディレクトリ
    temp_dir = tempfile.mkdtemp()

    # 各デバイスのスクリーンショットを配置
    current_row = 1
    max_screenshot_cols = 0
    screenshot_info = []

    for device in devices:
        screenshot_path = os.path.join(screenshots_dir, f"{screen_name}_{device}.png")

        if os.path.exists(screenshot_path):
            try:
                img = XLImage(screenshot_path)
                original_width = img.width
                original_height = img.height

                # デバイスに応じたスケール設定
                if device == 'PC':
                    scale = 0.5  # PCは50%
                elif device == 'タブレット':
                    scale = 0.6  # タブレットは60%
                else:  # スマホ
                    scale = 0.7  # スマホは70%

                target_width = int(original_width * scale)
                target_height = int(original_height * scale)
                screenshot_cols = (target_width // pixels_per_col) + 1

                if screenshot_cols > max_screenshot_cols:
                    max_screenshot_cols = screenshot_cols

                # ヘッダー行
                header_text = f"スクリーンショット（{device}）"
                ws.cell(row=current_row, column=1, value=header_text)
                header_cell = ws.cell(row=current_row, column=1)
                header_cell.fill = HEADER_FILL
                header_cell.font = HEADER_FONT
                header_cell.border = BORDER
                header_cell.alignment = CENTER_ALIGN

                # 画像配置行
                img_start_row = current_row + 1
                img.width = target_width
                img.height = target_height
                ws.add_image(img, f'A{img_start_row}')

                # 要素番号マーカーを配置
                device_positions = element_positions.get(device, {})
                for elem_no, (orig_x, orig_y) in device_positions.items():
                    if elem_no > len(elements):
                        continue

                    # スケール変換
                    scaled_x = int(orig_x * scale)
                    scaled_y = int(orig_y * scale)

                    # 番号マーカー画像を生成
                    img_path = os.path.join(temp_dir, f"marker_{device}_{elem_no}.png")
                    create_number_image(elem_no, img_path, marker_size)

                    marker_img = XLImage(img_path)
                    marker_img.width = marker_size
                    marker_img.height = marker_size

                    # セル位置を計算
                    col_offset = int(scaled_x // pixels_per_col)
                    row_offset = int(scaled_y // row_height_px)

                    target_col = 1 + col_offset
                    target_row = img_start_row + row_offset

                    cell_ref = f"{get_column_letter(target_col)}{target_row}"
                    ws.add_image(marker_img, cell_ref)

                # 必要な行数を計算
                rows_needed = (target_height // row_height_px) + 2

                screenshot_info.append({
                    'device': device,
                    'start_row': current_row,
                    'end_row': current_row + rows_needed,
                    'cols': screenshot_cols,
                    'scale': scale
                })

                print(f"  {device}: {original_width}x{original_height} -> {target_width}x{target_height}, マーカー: {len(device_positions)}個")

                current_row += rows_needed + 2  # 次のスクリーンショットとの間隔

            except Exception as e:
                print(f"警告: {device}スクリーンショット挿入エラー: {e}")
                current_row += 3
        else:
            print(f"警告: スクリーンショットが見つかりません: {screenshot_path}")

    # ヘッダーのマージ処理
    for info in screenshot_info:
        end_col = get_column_letter(max_screenshot_cols)
        ws.merge_cells(f'A{info["start_row"]}:{end_col}{info["start_row"]}')

    # 列幅設定
    for col in range(1, max_screenshot_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = col_width

    # 空白列
    ws.column_dimensions[get_column_letter(max_screenshot_cols + 1)].width = 2

    # 要素一覧（右側）
    start_col = max_screenshot_cols + 2
    headers = ["No", "要素名", "要素種別", "必須", "機能説明", "初期値", "バリデーション", "備考"]

    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=start_col + col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER_ALIGN

    for row_idx, elem in enumerate(elements, 2):
        row_data = [
            elem.get('no', row_idx - 1),
            elem.get('element_name', ''),
            elem.get('element_type', ''),
            elem.get('required', ''),
            elem.get('function', ''),
            elem.get('initial_value', ''),
            elem.get('validation', ''),
            elem.get('remarks', '')
        ]
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx, column=start_col + col_idx, value=value)
            cell.border = BORDER
            cell.alignment = CENTER_ALIGN if col_idx in [0, 2, 3] else LEFT_ALIGN

    # 要素一覧の列幅設定
    col_widths = [6, 25, 15, 6, 35, 15, 15, 20]
    for idx, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(start_col + idx)].width = width

    return ws


def create_element_sheet(wb, screen_name: str, elements: list[dict], screenshot_path: str = None):
    """要素シートを作成（左：スクリーンショット、右：要素一覧）- 単一スクリーンショット版"""
    ws = wb.create_sheet(title=screen_name)

    # スクリーンショットのサイズ設定
    screenshot_col_width = 2.5  # 各列の幅（文字数）

    # デフォルト値
    screenshot_scale = 1.0
    screenshot_cols = 60

    # スクリーンショット画像の挿入
    if screenshot_path and os.path.exists(screenshot_path):
        try:
            img = XLImage(screenshot_path)
            # 元画像の縦横比率を取得
            original_width = img.width
            original_height = img.height
            aspect_ratio = original_height / original_width if original_width > 0 else 0.75

            # スクリーンショットサイズを設定（元画像の60%）
            target_width = int(original_width * 0.6)
            target_height = int(target_width * aspect_ratio)
            screenshot_scale = target_width / original_width

            # スクリーンショット幅から必要な列数を逆算
            # Excelの1列幅2.5文字 ≈ 約17ピクセル（実測値に近い）
            pixels_per_col = 17
            screenshot_cols = (target_width // pixels_per_col) + 1

            img.width = target_width
            img.height = target_height
            ws.add_image(img, 'A2')
            print(f"スクリーンショット挿入: {screenshot_path} ({original_width}x{original_height} -> {target_width}x{target_height})")
            print(f"ヘッダー列数: {screenshot_cols} (A-{get_column_letter(screenshot_cols)})")
        except Exception as e:
            print(f"警告: スクリーンショット挿入エラー: {e}")
            ws.cell(row=2, column=1, value="（スクリーンショット未挿入）")
    else:
        print(f"警告: スクリーンショットが見つかりません: {screenshot_path}")

    # 要素番号マーカーをスクリーンショット上に配置
    import tempfile
    temp_dir = tempfile.mkdtemp()
    marker_size = 24

    # 要素の位置定義（元画像のピクセル座標）
    # フォーマット: (x, y) - 2400x1200の画像上での位置
    element_positions = {
        1: (150, 25),      # H001 - 画面タイトル
        2: (30, 25),       # H002 - 戻るボタン
        3: (950, 25),      # H003 - 編集リストラベル
        4: (1100, 25),     # H004 - 編集リスト選択
        5: (80, 85),       # N001 - 見積G一覧タブ
        6: (180, 85),      # N002 - 見積明細タブ
        7: (300, 85),      # N003 - 修理依頼一覧タブ
        8: (430, 85),      # N004 - 修理明細タブ
        9: (540, 85),      # N005 - メーカー保守一覧タブ
        10: (680, 85),     # N006 - 院内点検一覧タブ
        11: (100, 130),    # F001 - 見積区分フィルター
        12: (280, 130),    # F002 - 見積フェーズフィルター
        13: (450, 130),    # F003 - ステータスフィルター
        14: (400, 300),    # T001 - 見積G一覧テーブル
    }

    for elem_no, (orig_x, orig_y) in element_positions.items():
        if elem_no > len(elements):
            break

        # スケール変換（元画像座標 → Excel上の座標）
        # A2セルからの相対位置をピクセルで計算
        scaled_x = int(orig_x * screenshot_scale)
        scaled_y = int(orig_y * screenshot_scale)

        # 番号マーカー画像を生成
        img_path = os.path.join(temp_dir, f"marker_{elem_no}.png")
        create_number_image(elem_no, img_path, marker_size)

        # openpyxlのAnchorを使用して位置指定
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor, OneCellAnchor
        from openpyxl.utils.units import pixels_to_EMU

        marker_img = XLImage(img_path)
        marker_img.width = marker_size
        marker_img.height = marker_size

        # A2セルを基準にオフセット
        # 1列 ≈ 17ピクセル（実測値）
        col_width_px = 17
        row_height_px = 15  # デフォルト行高さ約15px

        # 列と行を計算
        col_offset = int(scaled_x // col_width_px)
        row_offset = int(scaled_y // row_height_px)

        target_col = 1 + col_offset  # A列 = 1
        target_row = 2 + row_offset  # 2行目から

        # 画像を配置
        cell_ref = f"{get_column_letter(target_col)}{target_row}"
        marker_img.anchor = cell_ref
        ws.add_image(marker_img, cell_ref)

    # 左側：スクリーンショット用スペースのヘッダー
    end_col_letter = get_column_letter(screenshot_cols)
    ws.merge_cells(f'A1:{end_col_letter}1')
    cell = ws.cell(row=1, column=1, value="スクリーンショット")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = CENTER_ALIGN

    # スクリーンショット用の列幅設定
    for col in range(1, screenshot_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = screenshot_col_width

    # 1列分の空白を追加（スクリーンショットと要素テーブルの間）
    ws.column_dimensions[get_column_letter(screenshot_cols + 1)].width = 2

    # 右側：要素一覧（スクリーンショット列 + 空白列の次から開始）
    start_col = screenshot_cols + 2
    headers = ["No", "要素名", "要素種別", "必須", "機能説明", "初期値", "バリデーション", "備考"]

    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=start_col + col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER_ALIGN

    for row_idx, elem in enumerate(elements, 2):
        row_data = [
            elem.get('no', row_idx - 1),
            elem.get('element_name', ''),
            elem.get('element_type', ''),
            elem.get('required', ''),
            elem.get('function', ''),
            elem.get('initial_value', ''),
            elem.get('validation', ''),
            elem.get('remarks', '')
        ]
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx, column=start_col + col_idx, value=value)
            cell.border = BORDER
            cell.alignment = CENTER_ALIGN if col_idx in [0, 2, 3] else LEFT_ALIGN

    # 要素一覧の列幅設定
    col_widths = [6, 20, 12, 6, 30, 12, 15, 15]
    for idx, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(start_col + idx)].width = width


def add_screen_to_index(ws, row: int, screen_name: str, file_path: str, summary: str = '', remarks: str = ''):
    """目次シートに画面を追加"""
    row_data = [row - 1, screen_name, file_path, summary, remarks]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = CENTER_ALIGN if col_idx == 1 else LEFT_ALIGN


def create_index_only_document(screens: list[dict], output_path: str):
    """
    目次とNo.画像シートのみを生成

    Args:
        screens: 画面情報のリスト [{'no': '1', 'name': '見積管理', 'path': 'app/...', 'summary': '...'}]
        output_path: 出力Excelファイルパス
    """
    wb = Workbook()

    # 目次シート作成
    index_ws = create_index_sheet(wb)

    # No.画像シート作成
    create_number_sheet(wb)

    for idx, screen in enumerate(screens, 2):
        screen_name = screen['name']
        file_path = screen.get('path', '')
        summary = screen.get('summary', '')

        # 目次に追加
        add_screen_to_index(index_ws, idx, screen_name, file_path, summary)

    # 保存
    wb.save(output_path)
    print(f"画面設計書（目次のみ）を生成しました: {output_path}")

    # サマリー出力
    print(f"\n=== 生成完了 ===")
    print(f"総画面数: {len(screens)}")


def create_design_document(screens: list[dict], elements_dir: str, screenshots_dir: str, output_path: str):
    """
    画面設計書を生成

    Args:
        screens: 画面情報のリスト [{'no': '1', 'name': '見積管理', 'path': 'app/...', 'summary': '...'}]
        elements_dir: 要素MDファイルが格納されているディレクトリ
        screenshots_dir: スクリーンショットが格納されているディレクトリ
        output_path: 出力Excelファイルパス
    """
    wb = Workbook()

    # 目次シート作成
    index_ws = create_index_sheet(wb)

    # No.画像シート作成
    create_number_sheet(wb)

    # 全画面の要素を収集
    all_elements = {}

    for idx, screen in enumerate(screens, 2):
        screen_name = screen['name']
        file_path = screen.get('path', '')
        summary = screen.get('summary', '')

        # 目次に追加
        add_screen_to_index(index_ws, idx, screen_name, file_path, summary)

        # 要素ファイル読み込み
        elements_file = os.path.join(elements_dir, f"{screen_name}_elements.md")
        elements = parse_elements_md(elements_file)
        all_elements[screen_name] = elements

        # スクリーンショットパス
        screenshot_path = os.path.join(screenshots_dir, f"{screen_name}.png")

        # 要素シート作成（スクリーンショット付き）
        create_element_sheet(wb, screen_name, elements, screenshot_path)

    # 保存
    wb.save(output_path)
    print(f"画面設計書を生成しました: {output_path}")

    # サマリー出力
    print(f"\n=== 生成完了 ===")
    print(f"総画面数: {len(screens)}")
    for screen_name, elements in all_elements.items():
        print(f"  - {screen_name}: {len(elements)}要素")


def parse_screen_list_md(file_path: str) -> list[dict]:
    """
    screen_list.mdファイルをパースして画面リストを返す

    期待フォーマット:
    | No | 画面名 | ファイルパス | 概要 |
    """
    screens = []

    if not os.path.exists(file_path):
        return screens

    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')
    table_lines = [line for line in lines if line.strip().startswith('|')]

    if len(table_lines) < 3:
        return screens

    for line in table_lines[2:]:
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        if len(cells) >= 4:
            screens.append({
                'no': cells[0],
                'name': cells[1],
                'path': cells[2],
                'summary': cells[3] if len(cells) > 3 else ''
            })

    return screens


def main():
    parser = argparse.ArgumentParser(description='画面設計書Excel生成（全画面統合版）')
    parser.add_argument('--output-dir', default='.', help='出力ディレクトリ')
    parser.add_argument('--elements-dir', default='./elements', help='要素MDファイルのディレクトリ')
    parser.add_argument('--screenshots-dir', default='./screenshots', help='スクリーンショットのディレクトリ')
    parser.add_argument('--screens', nargs='+', help='画面名リスト（例: 見積管理 資産マスタ）')
    parser.add_argument('--screen-list', help='screen_list.mdファイルパス')
    parser.add_argument('--index-only', action='store_true', help='目次とNo.画像シートのみ生成')

    args = parser.parse_args()

    # 出力ディレクトリ作成
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 画面リスト作成
    screens = []

    # screen_list.mdから読み込み
    if args.screen_list and os.path.exists(args.screen_list):
        screens = parse_screen_list_md(args.screen_list)
        print(f"screen_list.mdから{len(screens)}画面を読み込みました")
    elif args.screens:
        for idx, name in enumerate(args.screens, 1):
            screens.append({
                'no': str(idx),
                'name': name,
                'path': '',
                'summary': ''
            })
    else:
        # elements_dirから自動検出
        elements_dir = Path(args.elements_dir)
        if elements_dir.exists():
            for md_file in sorted(elements_dir.glob('*_elements.md')):
                screen_name = md_file.stem.replace('_elements', '')
                screens.append({
                    'no': str(len(screens)+1),
                    'name': screen_name,
                    'path': '',
                    'summary': ''
                })

    if not screens:
        print("警告: 画面が見つかりません。--screens オプションで画面名を指定してください。")
        print("例: python3 generate_excel.py --screens 見積管理 資産マスタ QR発行")
        return

    # Excel生成
    output_path = output_dir / "画面設計書.xlsx"

    if args.index_only:
        # 目次とNo.画像シートのみ生成
        create_index_only_document(screens, str(output_path))
    else:
        create_design_document(screens, args.elements_dir, args.screenshots_dir, str(output_path))


if __name__ == '__main__':
    main()
