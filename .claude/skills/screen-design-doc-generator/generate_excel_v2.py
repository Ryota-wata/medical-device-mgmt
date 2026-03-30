#!/usr/bin/env python3
"""
画面設計書Excel生成スクリプト v2（新11カラム対応）

SSはPillowで合成画像として生成し、Excelには1枚の画像として貼り付ける。
行間のズレ問題を根本的に回避する。

依存関係:
    pip install openpyxl Pillow
"""

import argparse
import json
import math
import os
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.worksheet.hyperlink import Hyperlink
except ImportError:
    print("openpyxlがインストールされていません: pip install openpyxl")
    exit(1)

from PIL import Image as PILImage, ImageDraw, ImageFont
PILImage.MAX_IMAGE_PIXELS = None  # fullPage SSの巨大画像を許可

# === スタイル定義 ===
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
MODAL_HEADER_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 新11カラム定義
COLUMNS_V2 = [
    "No", "要素名", "要素種別", "データ型", "桁数",
    "必須", "操作仕様", "初期値", "バリデーション",
    "エラーメッセージ", "権限"
]
COL_WIDTHS_V2 = [6, 22, 12, 10, 10, 6, 35, 12, 20, 25, 15]

# センタリングするカラムインデックス
CENTER_COLS_V2 = {0, 2, 3, 5}  # No, 要素種別, データ型, 必須

# === SS合成用定数 ===
SS_LABEL_HEIGHT = 32
SS_LABEL_BG = (68, 114, 196)      # #4472C4
SS_MODAL_LABEL_BG = (39, 174, 96)  # #27AE60
SS_LABEL_FG = (255, 255, 255)
SS_GAP = 8
SS_SCALES = {'PC': 0.5, 'タブレット': 0.6, 'スマホ': 0.7}

try:
    SS_FONT = ImageFont.truetype("/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc", 16)
except Exception:
    SS_FONT = ImageFont.load_default()


# === SS合成関数 ===

def _compose_device_screenshots(screen_name, devices, screenshots_dir,
                                 modal_name=None, label_bg=None):
    """デバイス別SSをラベル付きで1枚に合成して返す"""
    if label_bg is None:
        label_bg = SS_LABEL_BG

    sections = []
    max_width = 0

    for device in devices:
        scale = SS_SCALES.get(device, 0.5)
        if modal_name:
            fname = f"{screen_name}_{modal_name}_{device}.png"
        else:
            fname = f"{screen_name}_{device}.png"
        path = os.path.join(screenshots_dir, fname)
        if not os.path.exists(path):
            continue

        img = PILImage.open(path)
        tw = int(img.width * scale)
        th = int(img.height * scale)
        img = img.resize((tw, th), PILImage.LANCZOS)

        label_text = f"スクリーンショット（{device}）"
        if modal_name:
            label_text = f"スクリーンショット（{device} - {modal_name}）"

        sections.append((label_text, img, label_bg))
        if tw > max_width:
            max_width = tw

    if not sections:
        return None

    total_height = sum(SS_LABEL_HEIGHT + img.height + SS_GAP for _, img, _ in sections) - SS_GAP
    composite = PILImage.new('RGB', (max_width, total_height), (255, 255, 255))
    draw = ImageDraw.Draw(composite)
    y = 0

    for label_text, img, bg in sections:
        draw.rectangle([0, y, max_width, y + SS_LABEL_HEIGHT], fill=bg)
        draw.text((8, y + 6), label_text, fill=SS_LABEL_FG, font=SS_FONT)
        y += SS_LABEL_HEIGHT
        composite.paste(img, (0, y))
        y += img.height + SS_GAP

    return composite


def compose_all_screenshots(screen_name, devices, screenshots_dir, modals=None):
    """ベースSS + モーダルSSを全て1枚に合成"""
    if modals is None:
        modals = []

    parts = []

    # ベースSS
    base = _compose_device_screenshots(screen_name, devices, screenshots_dir)
    if base:
        parts.append(base)

    # モーダルSS
    for modal in modals:
        modal_name = modal.get('name', '')
        modal_img = _compose_device_screenshots(
            screen_name, devices, screenshots_dir,
            modal_name=modal_name, label_bg=SS_MODAL_LABEL_BG
        )
        if modal_img:
            parts.append(modal_img)

    if not parts:
        return None

    max_width = max(p.width for p in parts)
    section_gap = 16
    total_height = sum(p.height for p in parts) + section_gap * (len(parts) - 1)

    final = PILImage.new('RGB', (max_width, total_height), (255, 255, 255))
    y = 0
    for p in parts:
        final.paste(p, (0, y))
        y += p.height + section_gap

    return final


# === MD パーサー ===

def detect_columns_version(lines: list[str]) -> int:
    table_lines = [l for l in lines if l.strip().startswith('|')]
    if len(table_lines) < 1:
        return 1
    header = table_lines[0]
    cols = [c.strip() for c in header.split('|')[1:-1]]
    if len(cols) >= 11:
        return 2
    return 1


def parse_elements_md_v2(file_path: str) -> tuple[list[dict], int]:
    elements = []
    if not os.path.exists(file_path):
        return elements, 1

    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')
    version = detect_columns_version(lines)
    table_lines = [l for l in lines if l.strip().startswith('|')]

    if len(table_lines) < 3:
        return elements, version

    for line in table_lines[2:]:
        cells = [c.strip() for c in line.split('|')[1:-1]]

        if version == 2 and len(cells) >= 11:
            elements.append({
                'no': cells[0],
                'element_name': cells[1],
                'element_type': cells[2],
                'data_type': cells[3],
                'digits': cells[4],
                'required': cells[5],
                'operation': cells[6],
                'initial_value': cells[7],
                'validation': cells[8],
                'error_message': cells[9],
                'permission': cells[10],
            })
        elif len(cells) >= 5:
            elements.append({
                'no': cells[0],
                'element_name': cells[1],
                'element_type': cells[2],
                'data_type': '--',
                'digits': '--',
                'required': cells[3] if len(cells) > 3 else '',
                'operation': cells[4] if len(cells) > 4 else '',
                'initial_value': cells[5] if len(cells) > 5 else '',
                'validation': cells[6] if len(cells) > 6 else '',
                'error_message': '--',
                'permission': '--',
            })

    return elements, version


# === Excel生成 ===

def apply_header_style(ws, row, columns, fill=None):
    if fill is None:
        fill = HEADER_FILL
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER_ALIGN


def create_index_sheet(wb, screens):
    ws = wb.active
    ws.title = "目次"

    headers = ["No", "画面名", "使用目的", "使用端末", "画面モックURL"]
    apply_header_style(ws, 1, headers)

    col_widths = [6, 40, 55, 20, 55]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    link_font = Font(color="0563C1", underline="single", size=11)

    for row_idx, screen in enumerate(screens, 2):
        device_str = '/'.join(screen.get('devices', ['PC']))
        sheet_name = screen['name'][:31]
        purpose = screen.get('purpose', '')
        mock_url = f"https://ryota-wata.github.io/medical-device-mgmt{screen.get('path', '')}"

        # No
        cell_no = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        cell_no.border = BORDER
        cell_no.alignment = CENTER_ALIGN

        # 画面名（シートへのリンク）
        cell_name = ws.cell(row=row_idx, column=2, value=screen['name'])
        cell_name.hyperlink = Hyperlink(ref=cell_name.coordinate, location=f"'{sheet_name}'!A1", display=screen['name'])
        cell_name.font = link_font
        cell_name.border = BORDER
        cell_name.alignment = LEFT_ALIGN

        # 使用目的
        cell_purpose = ws.cell(row=row_idx, column=3, value=purpose)
        cell_purpose.border = BORDER
        cell_purpose.alignment = LEFT_ALIGN

        # 使用端末
        cell_dev = ws.cell(row=row_idx, column=4, value=device_str)
        cell_dev.border = BORDER
        cell_dev.alignment = CENTER_ALIGN

        # 画面モックURL（リンク付き）
        cell_url = ws.cell(row=row_idx, column=5, value=mock_url)
        cell_url.hyperlink = mock_url
        cell_url.font = link_font
        cell_url.border = BORDER
        cell_url.alignment = LEFT_ALIGN

    return ws


def _create_number_image(number, output_path: str, size: int = 30):
    img = PILImage.new('RGBA', (size, size), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)
    draw.ellipse([0, 0, size-1, size-1], fill=(231, 76, 60, 255))

    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", int(size * 0.5))
    except Exception:
        font = ImageFont.load_default()

    text = str(number)
    bbox = draw.textbbox((0, 0), text, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    x = (size - text_width) // 2
    y = (size - text_height) // 2 - 2
    draw.text((x, y), text, fill=(255, 255, 255, 255), font=font)

    img.save(output_path, 'PNG')
    return output_path


def create_number_sheet(wb, temp_dir: str = None):
    import tempfile
    ws = wb.create_sheet(title="No.画像")
    ws.cell(row=1, column=1, value="スクリーンショットに貼り付ける要素No.画像")
    ws.merge_cells('A1:J1')
    ws.cell(row=2, column=1, value="※ 必要な番号をコピーしてスクリーンショット上に配置してください")
    ws.merge_cells('A2:J2')

    if temp_dir is None:
        temp_dir = tempfile.mkdtemp()

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 5

    start_row = 4
    marker_size = 30
    for row in range(start_row, start_row + 3):
        ws.row_dimensions[row].height = 30

    for i in range(1, 31):
        row = start_row + ((i - 1) // 10)
        col = ((i - 1) % 10) + 1
        img_path = os.path.join(temp_dir, f"num_{i}.png")
        _create_number_image(i, img_path, marker_size)
        img = XLImage(img_path)
        img.width = marker_size
        img.height = marker_size
        ws.add_image(img, f"{get_column_letter(col)}{row}")

    return ws


def create_side_by_side_sheet_v2(wb, screen_name: str, elements: list[dict],
                                  screenshots_dir: str, devices: list[str] = None,
                                  positions_dir: str = None, modals: list[dict] = None):
    """
    左右分割レイアウト:
    - 左: Pillow合成済みSS画像（1枚）
    - 右: 要素一覧テーブル（11カラム）
    """
    import tempfile

    if devices is None:
        devices = ['PC', 'タブレット', 'スマホ']
    if modals is None:
        modals = []

    sheet_title = screen_name[:31]
    ws = wb.create_sheet(title=sheet_title)
    ws.sheet_format.defaultRowHeight = 15

    col_width = 2.5
    pixels_per_col = 17

    # === 左側: 合成SS画像を生成・配置 ===
    composite = compose_all_screenshots(screen_name, devices, screenshots_dir, modals)

    ss_cols = 0
    if composite:
        # Excel貼り付け用スケール
        excel_scale = 1.0
        display_width = int(composite.width * excel_scale)
        display_height = int(composite.height * excel_scale)
        ss_cols = (display_width // pixels_per_col) + 1

        # 一時ファイルに保存
        tmp_dir = tempfile.mkdtemp()
        composite_path = os.path.join(tmp_dir, f"{screen_name}_composite.png")
        composite.save(composite_path, 'PNG')

        img = XLImage(composite_path)
        img.width = display_width
        img.height = display_height
        ws.add_image(img, 'A1')

        rows_for_image = math.ceil(display_height * 9525 / 190500) + 1
        print(f"  合成SS: {composite.width}x{composite.height} -> {display_width}x{display_height} ({rows_for_image}行)")
    else:
        ss_cols = 30

    # 列幅設定（左側）
    for col in range(1, ss_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = col_width

    # 空白列
    ws.column_dimensions[get_column_letter(ss_cols + 1)].width = 2

    # === 右側: 目次リンク + 要素一覧テーブル ===
    start_col = ss_cols + 2

    # 目次に戻るリンク（row=1）
    back_cell = ws.cell(row=1, column=start_col, value="← 目次に戻る")
    back_cell.hyperlink = Hyperlink(ref=back_cell.coordinate, location="'目次'!A1", display="← 目次に戻る")
    back_cell.font = Font(color="0563C1", underline="single", size=10)

    # ベース画面要素（row=2から）
    base_elements = [e for e in elements if not e.get('no', '').startswith('M')]
    _write_elements_table(ws, 2, base_elements, start_col=start_col)

    # モーダル要素
    modal_elements = [e for e in elements if e.get('no', '').startswith('M')]
    if modal_elements:
        modal_start_row = len(base_elements) + 4
        ws.cell(row=modal_start_row, column=start_col, value="モーダル要素")
        ws.cell(row=modal_start_row, column=start_col).font = Font(bold=True, size=12)
        _write_elements_table(ws, modal_start_row + 1, modal_elements,
                              start_col=start_col, header_fill=MODAL_HEADER_FILL)

    # 右側列幅設定
    for idx, width in enumerate(COL_WIDTHS_V2):
        ws.column_dimensions[get_column_letter(start_col + idx)].width = width

    return ws


def _write_elements_table(ws, current_row, elements, start_col=1, header_fill=None):
    if header_fill is None:
        header_fill = HEADER_FILL

    for col_idx, header in enumerate(COLUMNS_V2):
        cell = ws.cell(row=current_row, column=start_col + col_idx, value=header)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER_ALIGN
    current_row += 1

    for elem in elements:
        row_data = [
            elem.get('no', ''),
            elem.get('element_name', ''),
            elem.get('element_type', ''),
            elem.get('data_type', '--'),
            elem.get('digits', '--'),
            elem.get('required', ''),
            elem.get('operation', ''),
            elem.get('initial_value', ''),
            elem.get('validation', ''),
            elem.get('error_message', '--'),
            elem.get('permission', '--'),
        ]
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=current_row, column=start_col + col_idx, value=value)
            cell.border = BORDER
            cell.alignment = CENTER_ALIGN if col_idx in CENTER_COLS_V2 else LEFT_ALIGN
        current_row += 1

    return current_row


def main():
    parser = argparse.ArgumentParser(description='画面設計書Excel生成 v2')
    parser.add_argument('--output-dir', default='.', help='出力ディレクトリ')
    parser.add_argument('--elements-dir', default='./elements', help='要素MDファイルのディレクトリ')
    parser.add_argument('--screenshots-dir', default='./screenshots', help='スクリーンショットのディレクトリ')
    parser.add_argument('--screens', nargs='+', help='画面名リスト')
    parser.add_argument('--screen-configs', help='screen_configs.jsonパス')

    args = parser.parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    screens = []
    if args.screen_configs and os.path.exists(args.screen_configs):
        with open(args.screen_configs, 'r', encoding='utf-8') as f:
            config = json.load(f)
        screens = config.get('screens', [])
    elif args.screens:
        for name in args.screens:
            screens.append({'name': name, 'devices': ['PC']})
    else:
        elements_dir = Path(args.elements_dir)
        if elements_dir.exists():
            for md_file in sorted(elements_dir.glob('*_elements.md')):
                screen_name = md_file.stem.replace('_elements', '')
                screens.append({'name': screen_name, 'devices': ['PC']})

    if not screens:
        print("警告: 画面が見つかりません。")
        return

    wb = Workbook()
    create_index_sheet(wb, screens)
    print("目次シート作成完了")
    create_number_sheet(wb)
    print("No.画像シート作成完了")

    success = fail = skipped = 0
    for idx, screen in enumerate(screens, 1):
        screen_name = screen['name']
        devices = screen.get('devices', ['PC'])
        modals = screen.get('modals', [])

        elements_file = os.path.join(args.elements_dir, f"{screen_name}_elements.md")
        if not os.path.exists(elements_file):
            alt_name = screen_name.replace('画面', '')
            alt_file = os.path.join(args.elements_dir, f"{alt_name}_elements.md")
            if os.path.exists(alt_file):
                elements_file = alt_file
            else:
                print(f"SKIP [{idx}]: {screen_name} - 要素一覧MDなし")
                skipped += 1
                continue

        elements, version = parse_elements_md_v2(elements_file)
        if not elements:
            print(f"SKIP [{idx}]: {screen_name} - 要素が0件")
            skipped += 1
            continue

        try:
            create_side_by_side_sheet_v2(
                wb, screen_name, elements,
                args.screenshots_dir, devices=devices, modals=modals
            )
            success += 1
            print(f"OK [{idx}/{len(screens)}]: {screen_name} ({len(elements)}要素, v{version})")
        except Exception as e:
            fail += 1
            print(f"FAIL [{idx}]: {screen_name} - {e}")
            import traceback
            traceback.print_exc()

    output_path = output_dir / "画面設計書.xlsx"
    wb.save(str(output_path))
    print(f"\n=== 完了 ===")
    print(f"成功: {success} / スキップ: {skipped} / 失敗: {fail} / 合計: {len(screens)}")
    print(f"出力: {output_path}")


if __name__ == '__main__':
    main()
